import os
import time

import openpyxl
from loguru import logger
from openpyxl import load_workbook
from pandas import DataFrame
from re import sub


class ExtractWorkbook2Sheets:
    def __init__(self, src: str, dst: str, save_intermediate: bool = True) -> None:
        self.src_file = src
        self.dst_folder = dst
        self.save_intermediate = save_intermediate
        self.tic = time.time()
        self.wb = None
        self.sheet = None
        self.subject_name = None

        self.sheets = {}

        os.makedirs(self.dst_folder, exist_ok=True)

    def __call__(self) -> dict:
        """Extract workbook to sheets"""
        # total_memory = psutil.virtual_memory().total / (1024**3)  # in GB
        # if total_memory < 30:  # warn if memory is less than 32GB
        #     raise MemoryError('Not enough memory to extract workbook to sheets, 32 GB of RAM is required')
        logger.info('Extract workbook to sheets is running...')
        if not self.src_file.endswith('.xlsx') and not os.path.exists(self.src_file):
            raise ValueError(f'{self.src_file} is not a valid ".xlsx" file')
        self.extract_sheets()

        return self.sheets

    def __del__(self) -> None:
        """What time is it"""
        logger.info(f'Execution time: {round((time.time() - self.tic) / 60, 1)} minutes')

    def extract_sheets(self) -> None:
        """Extract sheets"""
        wb = self.load_file()  # load workbook

        for sheet_name in wb.sheetnames:  # loop through sheets
            if self.check_sheet_name(sheet_name):
                old_sheet = wb[sheet_name]  # extract sheet
                clean_sheet_name = self.get_clean_sheet_name(sheet_name)

                if self.save_intermediate:
                    new_wb = openpyxl.Workbook()  # create new workbook
                    new_sheet = new_wb.active  # get active sheet
                    new_sheet.title = clean_sheet_name
                    for row in old_sheet:  # copy old sheet to new sheet
                        for cell in row:
                            new_sheet[cell.coordinate].value = cell.value
                    new_wb.save(f'{os.path.join(self.dst_folder, clean_sheet_name)}.xlsx')
                    new_wb.close()

                else:  # store in dict instead of saving files
                    self.sheets[clean_sheet_name] = DataFrame(old_sheet.values)

    @staticmethod
    def get_clean_sheet_name(sheet_name: str) -> str:
        """Get a clean sheet name"""
        if '_' in sheet_name:
            sheet_name = sheet_name.split('_')[1]  # get sheet name after '_'
            sheet_name = sub("[^0-9]", "", sheet_name)  # remove non-numberic characters
            return sheet_name
        if ' ' in sheet_name:
            sheet_name = sheet_name.split(' ')[1]  # get sheet name after '_'
            sheet_name = sub("[^0-9]", "", sheet_name)  # remove non-numberic characters
            return sheet_name
        return f'fix_me_{sheet_name}'

    @staticmethod
    def check_sheet_name(sheet_name: str) -> bool:
        """Check sheet name"""
        if ('_' in sheet_name or '2 ' in sheet_name) and '#' not in sheet_name:
            logger.info(f'Extract sheet -> {sheet_name}')
            return True
        logger.info(f'Ignore sheet -> {sheet_name}')
        return False

    def load_file(self) -> openpyxl.Workbook or None:
        """Load file"""
        if not self.src_file.startswith('.'):  # avoid loading hidden tmp file
            self.subject_name = self.src_file.strip('.xlsx')
            return load_workbook(self.src_file, read_only=True, data_only=True, keep_vba=False, keep_links=False)
        return None
